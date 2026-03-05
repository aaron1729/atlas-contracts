"""
Per-cohort player statistics:
  - Top 5 words
  - Most unique contract (lowest P under independence model)
  - Most prototypical contract (highest P under independence model)
Also: overall top 10 words, most unique / prototypical overall.
"""

import csv
import re
from collections import Counter

import openpyxl

import clean

# ── Global word probability (independence model) ──────────────────────────────

contracts_3 = [c for c in clean.translated_contracts if len(c) == 3]
words_global = [w for c in contracts_3 for w in c]
total_words  = len(words_global)
word_counts  = Counter(words_global)
word_prob    = {w: n / total_words for w, n in word_counts.items()}


def contract_p(contract):
    """P(contract) under independence using global word distribution."""
    p = 1.0
    for w in contract:
        p *= word_prob.get(w, 1 / (total_words + 1))
    return p


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_non_player(cells):
    """True if any cell marks this row as a coach or captain."""
    keywords = ("captain", "coach", "co-captain")
    for cell in cells:
        if any(cell.lower().startswith(k) for k in keywords):
            return True
    return False


def _player_contracts(ws, contract_col_name):
    """
    Yield translated contract tuples for player rows in a roster sheet.
    Handles sheets that have a 'Role' column and those that don't.
    Skips blank rows, header rows, and coach/captain rows.
    """
    rows = [
        [str(c).strip() if c is not None else "" for c in r]
        for r in ws.iter_rows(values_only=True)
    ]
    hi = next(
        (i for i, row in enumerate(rows) if contract_col_name in row), None
    )
    if hi is None:
        return
    header   = rows[hi]
    ci       = header.index(contract_col_name)
    role_col = next(
        (header.index(h) for h in ("Role",) if h in header), None
    )
    for row in rows[hi + 1:]:
        if len(row) <= ci:
            continue
        raw = row[ci].strip()
        if not raw:
            continue
        if role_col is not None:
            role = row[role_col].lower() if len(row) > role_col else ""
            if role and role not in ("student", "player", ""):
                continue
        elif _is_non_player(row):
            continue
        c = clean._parse_contract(raw)
        if not c or len(c) < 2:
            continue
        t = clean._translate(c)
        if t and len(t) >= 2:
            yield t


def _sf15_player_contracts():
    """SF15 Roster: only 'Player N' rows (not Captain or Coach N)."""
    ws = _wb["SF15 Roster"]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells or not cells[0].startswith("Player"):
            continue
        if len(cells) > 3:
            c = clean._parse_contract(cells[3])
            if c:
                t = clean._translate(c)
                if t:
                    yield t


# ── Data collection ───────────────────────────────────────────────────────────

_wb = openpyxl.load_workbook(
    "data/Rainy Day Notes Sign Ups.xlsx", read_only=True, data_only=True
)

COHORTS = [
    ("LAS7 Roster",   "Contract",  "LAS7"),
    ("LAS8 Roster",   "Contract",  "LAS8"),
    ("LAS10 Roster",  "Contract",  "LAS10"),
    ("LAS11 Roster",  "Contract",  "LAS11"),
    ("LAS12 Roster",  "Contract",  "LAS12"),
    ("ATX10 Roster",  "Contract",  "ATX10"),
    ("ATX11 Roster",  "Najiba",    "ATX11"),
    ("ATX12 Roster",  "Contract",  "ATX12"),
    ("ATX13 Roster",  "Contract",  "ATX13"),
    ("ATX14 Roster",  "Contract",  "ATX14"),
    ("ATX15 Roster",  "Contract",  "ATX15"),
]

cohort_data = {}  # label → {"players": [...], "top5": [...], "unique": ..., "proto": ...}

for sheet_name, col_name, label in COHORTS:
    players = [c for c in _player_contracts(_wb[sheet_name], col_name) if len(c) == 3]
    if not players:
        continue
    cw = Counter(w for c in players for w in c)
    cohort_data[label] = {
        "n":      len(players),
        "top5":   cw.most_common(5),
        "unique": min(players, key=contract_p),
        "proto":  max(players, key=contract_p),
    }

# ATX16: complete roster is in contracts - Sheet1.csv (xlsx roster is missing one player)
_atx16 = []
with open("data/contracts - Sheet1.csv", newline="", encoding="utf-8") as _f:
    for _row in csv.DictReader(_f):
        if _row.get("cohort", "").strip() != "ATX16":
            continue
        adjs = tuple(
            re.sub(r"[^\w-]", "", _row[k]).strip().lower()
            for k in ("adjective 1", "adjective 2", "adjective 3")
        )
        adjs = tuple(a for a in adjs if a)
        t = clean._translate(adjs)
        if t and len(t) == 3:
            _atx16.append(t)
if _atx16:
    cw = Counter(w for c in _atx16 for w in c)
    cohort_data["ATX16"] = {
        "n":      len(_atx16),
        "top5":   cw.most_common(5),
        "unique": min(_atx16, key=contract_p),
        "proto":  max(_atx16, key=contract_p),
    }


# ── Print ─────────────────────────────────────────────────────────────────────

print("=== Top 10 words overall ===")
for w, n in word_counts.most_common(10):
    print(f"  {w}: {n}")

print()
print("=== Top 5 words per cohort (player roster) ===")
print(f"{'Cohort':<8}  {'Players':>7}  Top 5 words")
print("-" * 70)
for label, d in cohort_data.items():
    top5_str = ", ".join(f"{w} ({n})" for w, n in d["top5"])
    print(f"{label:<8}  {d['n']:>7}  {top5_str}")

print()
print("=== Most unique player contract per cohort (lowest P under independence) ===")
print(f"{'Cohort':<8}  {'Contract':<50}  {'P':>12}")
print("-" * 75)
for label, d in cohort_data.items():
    c = d["unique"]
    p = contract_p(c)
    print(f"{label:<8}  {str(c):<50}  {p:.2e}")

print()
print("=== Most prototypical player contract per cohort (highest P under independence) ===")
print(f"{'Cohort':<8}  {'Contract':<50}  {'P':>12}")
print("-" * 75)
for label, d in cohort_data.items():
    c = d["proto"]
    p = contract_p(c)
    print(f"{label:<8}  {str(c):<50}  {p:.2e}")

print()
print("=== Most unique contract overall (among all 3-word translated contracts) ===")
most_unique = min(contracts_3, key=contract_p)
print(f"  {most_unique}  P = {contract_p(most_unique):.2e}")

print()
print("=== Most prototypical contract overall ===")
most_proto = max(contracts_3, key=contract_p)
print(f"  {most_proto}  P = {contract_p(most_proto):.2e}")
print(f"  (independent model peak: loving × loving × loving would be "
      f"P = {word_prob['loving']**3:.4%}, but no contract repeats a word)")
