import csv
import os
import re
from collections import Counter

import openpyxl


# ── Shared contract parser ────────────────────────────────────────────────────

def _parse_contract(raw):
    """Extract a tuple of lowercase adjectives from any contract string variant.

    Handles:
    - Full sentences: "I am (a/an)? ... leader" (optional article, any case)
    - Plain comma-separated: "Powerful, Joyful, Kind"
    - Space-separated (no commas): "Brave Powerful Free"
    - "Leader!" suffix: "Loving, Kind, Brave Leader!"
    - Bilingual slash: "Peaceful, loving, abundant / pacífico, amado, abundante"
    - "and" / "&" separators: "loving, kind and compassionate"
    - Period-as-separator: "Trusting. loving, worthy"
    - Space-separated after "and" normalization: "Trusting Powerful and Giving"
    """
    raw = raw.strip()
    if not raw:
        return None

    # Full-sentence format: "I am (a/an)? ... leader"
    m = re.search(r"(?i)i\s+am\s+(?:an?\s+)?(.*?)\s*leader", raw)
    if m:
        middle = m.group(1).strip()
    else:
        middle = raw
        # Bilingual: take only the part before "/"
        if "/" in middle:
            middle = middle.split("/")[0].strip()
        # Strip trailing "leader" with optional punctuation
        middle = re.sub(r"(?i)\s*leader[!.]?\s*$", "", middle).strip()

    # Treat period-after-word as a comma (e.g. "Trusting. loving, worthy")
    middle = re.sub(r"(?<=\w)\.\s*", ",", middle)
    # Remove punctuation except word chars, hyphens, spaces, commas, ampersands
    middle = re.sub(r"[^\w\s,&-]", "", middle)
    # Normalize "and" and "&" to commas
    middle = re.sub(r"\s+and\s+", ",", middle, flags=re.IGNORECASE)
    middle = re.sub(r"\s*&\s*", ",", middle)

    # Split by comma, then split any part that still has spaces
    # (handles both "Brave Powerful Free" and "Trusting Powerful, Giving")
    parts = []
    for chunk in middle.split(","):
        parts.extend(chunk.split())

    adjectives = tuple(re.sub(r"[^\w-]", "", p).lower() for p in parts if p)
    return adjectives if adjectives else None


# ── xlsx helpers ──────────────────────────────────────────────────────────────

_wb = openpyxl.load_workbook(
    "data/Rainy Day Notes Sign Ups.xlsx", read_only=True, data_only=True
)


def _col_xlsx(sheet_name, col_name):
    """Yield contract tuples from a named column in an xlsx sheet.

    Scans all rows to find the first one containing col_name as a cell value
    (handles metadata rows before the real header, like ATX13). Sub-header rows
    mid-sheet (where the contract cell equals a column name) produce 1-element
    tuples and are filtered out by the len >= 2 requirement.
    """
    ws = _wb[sheet_name]
    rows = [
        [str(c).strip() if c is not None else "" for c in r]
        for r in ws.iter_rows(values_only=True)
    ]
    hi = next((i for i, row in enumerate(rows) if col_name in row), None)
    if hi is None:
        return
    ci = rows[hi].index(col_name)
    for row in rows[hi + 1:]:
        if len(row) <= ci:
            continue
        c = _parse_contract(row[ci])
        if c and len(c) >= 2:
            yield c


def _signup_xlsx(sheet_name):
    """Yield contracts from sign-up sheets (TRUE/FALSE in col 0, contract in col 5).

    openpyxl returns Excel booleans as Python True/False; str(True) == "True",
    which uppercases to "TRUE". Metadata rows, header rows, and blank rows all
    have non-boolean values in col 0 and are naturally skipped.
    """
    ws = _wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) < 6 or cells[0].upper() not in ("TRUE", "FALSE"):
            continue
        name, contract_raw = cells[1], cells[5]
        if not name or not contract_raw:
            continue
        c = _parse_contract(contract_raw)
        if c and len(c) >= 2:
            yield c


def _sf15_xlsx():
    """Yield contracts from SF15 Roster (data rows identified by col-0 marker)."""
    ws = _wb["SF15 Roster"]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not cells:
            continue
        marker = cells[0]
        if not (
            marker.startswith("Player")
            or marker == "Captain"
            or marker.startswith("Coach")
        ):
            continue
        if len(cells) > 3:
            c = _parse_contract(cells[3])
            if c and len(c) >= 2:
                yield c


# ── Guard: error on unrecognized data files ───────────────────────────────────

KNOWN_FILES = {
    "Rainy Day Notes Sign Ups.xlsx",                              # master xlsx
    "contracts - Sheet1.csv",                                     # separate; not in the xlsx
    "LAS11 Roster RD - Sheet1.csv",                               # old CSV, superseded by xlsx
    "ZZZ old individual csv files from Rainy Day Notes xlsx",     # folder of archived old CSVs
}

data_files = set(f for f in os.listdir("data") if not f.startswith("."))
unknown = data_files - KNOWN_FILES
if unknown:
    raise ValueError(f"Unhandled data file(s) in data/: {unknown}")


# ── Rainy Day Notes Sign Ups.xlsx ─────────────────────────────────────────────
#
# Sheets skipped (no contract data — they track who committed to writing notes):
#   (TEMPLATE), RDN Instructions, RDN Mailbox instructions!, Sheet55,
#   ATX16/ATX15/LAS11/LAS12/LAS10/ATX14 Rainy Day Note Sign Up,
#   ATX9/ATX10/LAS8/LAS7/ATX13 Sign Up, LAS8 Sign Up
#
# ATX11 and ATX12 Sign Up DO have contracts (TRUE/FALSE + contract format).
# All other sign-up-style sheets only track set counts, not contracts.

# Rosters with "Contract" column
las12_contracts    = list(_col_xlsx("LAS12 Roster",  "Contract"))
las11_contracts    = list(_col_xlsx("LAS11 Roster",  "Contract"))
las10_contracts    = list(_col_xlsx("LAS10 Roster",  "Contract"))
las8_contracts     = list(_col_xlsx("LAS8 Roster",   "Contract"))
las7_contracts     = list(_col_xlsx("LAS7 Roster",   "Contract"))
las6_contracts     = list(_col_xlsx("LAS6 Roster",   "Contract"))
atx16_contracts    = list(_col_xlsx("ATX16 Roster",  "Contract"))
atx15_contracts    = list(_col_xlsx("ATX15 Roster",  "Contract"))
atx14_contracts    = list(_col_xlsx("ATX14 Roster",  "Contract"))
atx13_contracts    = list(_col_xlsx("ATX13 Roster",  "Contract"))  # metadata rows before header
atx12_roster       = list(_col_xlsx("ATX12 Roster",  "Contract"))
atx10_contracts    = list(_col_xlsx("ATX10 Roster",  "Contract"))  # one bilingual slash entry

# Rosters with "CONTRACT" column (all-caps header)
las5_contracts     = list(_col_xlsx("LAS5 Roster",   "CONTRACT"))
atx8_contracts     = list(_col_xlsx("ATX8 Roster",   "CONTRACT"))
atx3_contracts     = list(_col_xlsx("ATX3 Roster",   "CONTRACT"))

# Rosters with "I AM..." column
sf14_contracts     = list(_col_xlsx("SF14 Roster",   "I AM..."))
atx2_contracts     = list(_col_xlsx("ATX2 Roster",   "I AM..."))

# ATX11 Roster: two sections; contract col is mislabeled "Najiba" in the first header
atx11_roster       = list(_col_xlsx("ATX11 Roster",  "Najiba"))

# SF17: "Contract " header has trailing space — _col_xlsx strips cell values, so "Contract" matches
sf17_contracts     = list(_col_xlsx("SF17",          "Contract"))

# SF15: many blank/metadata rows before data; rows identified by "Player N"/"Captain"/"Coach N"
sf15_contracts     = list(_sf15_xlsx())

# Sign-up sheets with TRUE/FALSE col 0 and contract in col 5
atx12_signup       = list(_signup_xlsx("ATX12 Sign Up"))
atx11_signup       = list(_signup_xlsx("ATX11 Sign Up"))


# ── contracts - Sheet1.csv ────────────────────────────────────────────────────
# Pre-split into dedicated "adjective 1", "adjective 2", "adjective 3" columns.

sheet1_contracts = []
with open("data/contracts - Sheet1.csv", newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        adjs = tuple(
            re.sub(r"[^\w-]", "", row[k]).strip().lower()
            for k in ("adjective 1", "adjective 2", "adjective 3")
        )
        adjs = tuple(a for a in adjs if a)
        if adjs:
            sheet1_contracts.append(adjs)


# ── All contracts ─────────────────────────────────────────────────────────────

all_contracts = [
    *las12_contracts, *las11_contracts, *las10_contracts,
    *las8_contracts,  *las7_contracts,  *las6_contracts,  *las5_contracts,
    *atx16_contracts, *atx15_contracts, *atx14_contracts, *atx13_contracts,
    *atx12_roster,    *atx12_signup,
    *atx11_roster,    *atx11_signup,
    *atx10_contracts, *atx8_contracts,  *atx3_contracts,  *atx2_contracts,
    *sf17_contracts,  *sf15_contracts,  *sf14_contracts,
    *sheet1_contracts,
]

# ── Translations ─────────────────────────────────────────────────────────────
# Maps non-canonical adjective spellings to their canonical form.
# A value of None means "drop this word entirely".

TRANSLATIONS = {
    # Alternative forms of the same word
    "joyous":         "joyful",           # archaic/poetic alternative

    # Typos — near-certain
    "compasionate":   "compassionate",
    "compassionte":   "compassionate",
    "courageious":    "courageous",
    "couragous":      "courageous",
    "intiuitive":     "intuitive",
    "integreous":     "integrous",
    "determine":      "determined",       # wrong word form (verb/noun → adjective)
    "compassionte":   "compassionate",
    "orgasmin":       "orgasmic",
    "firece":         "fierce",
    "nuturing":       "nurturing",
    "unifing":        "unifying",
    "commited":       "committed",

    # Typos — possibly intentional / phonetic
    "power":          "powerful",         # noun; almost certainly meant the adjective
    "greatful":       "grateful",         # common phonetic misspelling, though possibly an intentional playful alternative
    "devine":         "divine",
    "woth":           "worthy",
    "acepting":       "accepting",

    # Creative / non-standard spellings
    "trans4mational": "transformational", # intentional leet-style
    "heartled":       "heart-led",        # missing hyphen

    # Spanish (one entry: Zerethsha Larios, ATX11)
    "creativa":       "creative",
    "apasionada":     "passionate",
    "valorada":       "valued",

    # Parsing artifact: "Very Loving" split into two words by the space-splitter
    # Also handled at contract level in CONTRACT_OVERRIDES below
    "very":           None,               # drop — not a standalone adjective here
}


def _translate(contract):
    result = tuple(
        TRANSLATIONS.get(adj, adj)
        for adj in contract
        if TRANSLATIONS.get(adj, adj) is not None
    )
    return result if result else None


# ── Contract-level overrides ───────────────────────────────────────────────────
# Fixes contracts that don't parse to exactly 3 words.
# A value of None means "drop this contract entirely".
# Applied after parsing, before TRANSLATIONS.
# Any non-3-word contract not listed here raises a ValueError.

CONTRACT_OVERRIDES = {
    # Parsing artifact: "Light Shining" (no hyphen) split into two words
    ('loving', 'empowering', 'light', 'shining'):     ('loving', 'empowering', 'light-shining'),

    # Parsing artifact: "Very Loving" split into two words
    ('joyful', 'honorable', 'very', 'loving'):        ('joyful', 'honorable', 'loving'),

    # Incomplete contracts (2 adjectives) — kept as-is (flagged in output)
    ('compassionate', 'loving'):                       ('compassionate', 'loving'),
    ('strong', 'joyful'):                              ('strong', 'joyful'),

    # 4-adjective contracts — kept as-is (flagged in output)
    # Gia Randazzo, ATX2
    ('powerful', 'courageous', 'forgiving', 'loving'): ('powerful', 'courageous', 'forgiving', 'loving'),
    # Laura Saavedra, ATX11 & ATX12 (same person, two cohorts; typo in ATX12)
    ('powerful', 'courageous', 'trusting', 'loving'):  ('powerful', 'courageous', 'trusting', 'loving'),
    ('powerful', 'couragous', 'trusting', 'loving'):   ('powerful', 'couragous', 'trusting', 'loving'),
}


def _apply_overrides(contracts):
    result = []
    for c in contracts:
        if len(c) in (2, 3):
            result.append(c)
        elif c in CONTRACT_OVERRIDES:
            override = CONTRACT_OVERRIDES[c]
            if override is not None:
                result.append(override)
        else:
            raise ValueError(f"Unhandled non-standard contract: {c!r}")
    return result


# ── Top adjectives ────────────────────────────────────────────────────────────

all_contracts = _apply_overrides(all_contracts)

all_adjectives = [adj for contract in all_contracts for adj in contract]
counts_raw = Counter(all_adjectives)

translated_contracts = [c for c in (_translate(ct) for ct in all_contracts) if c]
counts = Counter(adj for ct in translated_contracts for adj in ct)

non3 = [c for c in all_contracts if len(c) != 3]

if __name__ == "__main__":
    print("=== before translations ===")
    for adj, n in counts_raw.most_common():
        print(f"{adj}: {n}")
    print(f"total unique words: {len(counts_raw)}")
    print(f"total count of words: {sum(v for v in counts_raw.values())}")

    print("\n=== after translations ===")
    for adj, n in counts.most_common():
        print(f"{adj}: {n}")
    print(f"total unique words: {len(counts)}")
    print(f"total count of words: {sum(v for v in counts.values())}")

    if non3:
        print("\n=== non-3-word contracts (kept in statistics) ===")
        for c in non3:
            print(c)